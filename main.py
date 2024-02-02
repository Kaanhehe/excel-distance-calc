import sys
import os
from tracemalloc import start
from matplotlib import table
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import string

### VARIABLES ###
try:
    api_key = sys.argv[1]
    mode = sys.argv[2]
    if mode not in ["1", "2"]:
        print("Error: Invalid mode.")
        exit(1)
    if mode == "1":
        starting_point = sys.argv[3]
    if mode == "2":
        starting_point_adress_column = sys.argv[3]
    adress_column = sys.argv[4]
    distance_column = sys.argv[5]
    file_name = sys.argv[6]
    sheet_name = sys.argv[7]
except IndexError:
    api_key = input("Enter Google Maps API key: ")
    print("Mode 1: Enter the starting point. Mode 2: Take the starting point from the file.")
    mode = input("Do you want to use Mode 1 or Mode 2? (1/2): ")
    if mode not in ["1", "2"]:
        print("Error: Invalid mode.")
        exit(1)
    if mode == "1":
        starting_point = input("Enter starting point: ")
    if mode == "2":
        starting_point_adress_column = input("Enter starting point address column (A, B, C...): ")
    adress_column = input("Enter address column (A, B, C...): ")
    distance_column = input("Enter distance column(A, B, C...): ")
    file_name = input("Enter file name(without .xlsx): ") + ".xlsx" or "Mappe1.xlsx"
    sheet_name = input("Enter sheet name (or leave it empty): ") or "Tabelle1"

### MAIN ###

# Check if the file exists
if not os.path.isfile(file_name):
    print(f"Error: '{file_name}' file does not exist.")
    exit(1)

# Read the Excel file
df = pd.read_excel(file_name, sheet_name=sheet_name)
wb = load_workbook(file_name)
ws = wb.active

# Check if the API key is valid
response = requests.get(f"https://maps.googleapis.com/maps/api/distancematrix/json?units=imperial&origins=Seattle&destinations=San+Francisco&key={api_key}")
if response.status_code != 200:
    print("Error: Invalid API key.")
    exit(1)

def calculate_distance(starting_point, destination_addresses):
    url = "https://maps.googleapis.com/maps/api/distancematrix/json"

    destination_addresses = [address for address in destination_addresses if not isinstance(address, float)]
    destination_addresses = [address for address in destination_addresses if address]
    params = {
        "origins": starting_point,
        "destinations": "|".join(destination_addresses),
        "key": api_key
    }

    response = requests.get(url, params=params)
    data = response.json()

    distances = []
    for i, row in enumerate(data["rows"]):
        for j, element in enumerate(row["elements"]):
            distance = element["distance"]["text"].replace(" km", "")
            print(f"Distance from {starting_point} to {destination_addresses[i]}: {distance}")
            distances.append(distance)
    return distances[0] if distances else None

# Convert the letter of the column to its corresponding index
column_index = string.ascii_uppercase.index(adress_column.upper())
distance_index = string.ascii_uppercase.index(distance_column.upper())

# Calculate the distance for each row
### I get some warning here, but it works fine so idk, I'll just ignore it
if mode == "1":
    df[distance_index] = df.apply(lambda row: pd.Series([calculate_distance(starting_point, [row[column_index]])]), axis=1) # type: ignore
if mode == "2":
    starting_index = string.ascii_uppercase.index(starting_point_adress_column.upper()) # type: ignore
    df[distance_index] = df.apply(lambda row: pd.Series([calculate_distance([row[starting_index]], [row[column_index]])]), axis=1) # type: ignore
# Insert the calculated distance into the Distance column
for cell in ws[distance_column]: # type: ignore
    if cell.row > 1 and cell.value is None:
        if cell.row > 1 and cell.row - 1 < len(df):
            cell.value = df.iloc[cell.row - 3, df.columns.get_loc(distance_index)]

wb.save(file_name)

print("Done!!!")